# /gpfs/home/yuk/anaconda3/envs/py36/bin/python3
# -*- coding: utf-8 -*-

# filepath: /gpfs/home/yuk/LAI/out_inter/GLASS/month    LAI data
# calculate area : /gpfs/home/yuk/merged/veg_mean0.2.tif
import gdal
import numpy
import os
from osgeo import gdal, gdalconst
from osgeo import ogr, osr
import math
import datetime
import openpyxl
from scipy.interpolate import interp1d
from scipy.signal import savgol_filter
import matplotlib.pyplot as plt
import copyreg
import types
import multiprocess
from multiprocess.pool import ThreadPool as TPool
from tqdm import tqdm



filepath = r'/month'
ffile_list = os.listdir(filepath)
ffile_list.sort()
# lists1 = [[] for _ in range(72)]
# lists2 = [[] for _ in range(72)]
year = '2000'
day = []
image = numpy.zeros((3600, 7200, 46), dtype='float32')

l1 = 0
l2 = 0
for ff in ffile_list:
    date = (str(ff).split('_')[0])[0:4]
    if date == year and str(ff)[-1] == 'f':
        globe_data = gdal.Open(os.path.join(filepath, ff))
        im_width0 = globe_data.RasterXSize
        im_height0 = globe_data.RasterYSize  #
        image[:, :, l1] = globe_data.ReadAsArray(0, 0, im_width0, im_height0)
        day.append((str(ff).split('_')[0])[4:])
        l1=l1+1
    else:
        continue

threefile = []
date = []

for layer in range(0, len(day)):
    threefile.append(image[:,:,layer])
    # dif = datetime.date(int(year), int(str(day[layer])[0:2]), int(str(day[layer])[2:4])).__sub__(
    #     datetime.date(int(year), 1, 1)).days
    date.append(int(day[layer]))

file_ex = r'/veg_mean0.2.tif'
ex_data = gdal.Open(file_ex)
im_width0 = int(ex_data.RasterXSize)
im_height0 = int(ex_data.RasterYSize)
im_veg = ex_data.ReadAsArray(0, 0, im_width0, im_height0)
im_geotrans0 = ex_data.GetGeoTransform()
im_proj0 = ex_data.GetProjection()

# optima = numpy.full((im_height0, im_width0), numpy.nan, dtype='float16')
# outpath = r'D:\airtemp\fuse'
gs = numpy.full((im_height0, im_width0, 3), numpy.nan, dtype='float32')
for i in range(0, im_width0):
    for j in range(0, im_height0):
        if ~numpy.isnan(im_veg[i, j]) and im_veg[i, j] > 0:

            time = []
            lai = []
            for k in range(0, len(threefile)):
                if ~numpy.isnan(numpy.array(threefile[k])[i, j]) and numpy.array(threefile[k])[i, j] >= 0:
                    time.append(date[k])
                    lai.append(numpy.array(threefile[k])[i, j])
                    yy2.cell(k+1, 1).value = date[k]
                    yy2.cell(k+1, 2).value = numpy.array(threefile[k])[i, j]
                else:
                    continue
            if len(lai) >= 15:
                f = interp1d(time, lai, kind='cubic', fill_value='extrapolate')
                xx = numpy.linspace(1, 365, 365)
                yy = f(xx)
                smothyy = yy
                smoother_ts = savgol_filter(smothyy, window_length=21, polyorder=4)
                y_end = smoother_ts
                for kk in range(0,len(smoother_ts)):
                    yy2.cell(kk+1, 3).value = xx[kk]
                    yy2.cell(kk+1, 4).value = smoother_ts[kk]

                kkk = (y_end[200] - y_end[0]) / (xx[200] - xx[0])
                k11 = "%.8f" % kkk
                if kkk >= 0 or (str(k11).split(".")[1])[0:5] == '00000':
                    # y_end3 = y_end[numpy.argsort(y_end[:])]
                    y_end_zzong = y_end[int(list(xx).index(time[0])):365]
                    xx_zzong = xx[int(list(xx).index(time[0])):365]
                    y_end4 = y_end_zzong[numpy.argsort(y_end_zzong)]
                    # lommin = int(list(xx).index(time[0])) + list(yy_qian).index(numpy.nanmin(list(yy_qian)))
                    lomax = int(list(xx).index(time[0])) + list(y_end_zzong).index(numpy.nanmax(list(y_end_zzong)))
                    yy_qian = y_end[int(list(xx).index(time[0])):250]
                    yy_qian_paixu = yy_qian[numpy.argsort(yy_qian)]
                    yy_qian_paixu = yy_qian_paixu[yy_qian_paixu != 0]
                    # print(yy_qian_paixu)
                    y_end1_shi1 = numpy.nanmean(yy_qian_paixu[0:math.floor(len(yy_qian_paixu) * 0.1)])
                    y_end1_mo = numpy.nanmean(y_end4[math.floor(len(y_end4) * 0.9):365])
                    amplitude1 = y_end1_mo - y_end1_shi1
                    # amplitude2 = y_end1_mo - y_end1_shi2
                    amp1 = amplitude1 * 0.2 + y_end1_shi1
                    # print(y_end1_mo, y_end1_shi1, amp1)
                    if amp1 <= 0.2:
                        amp1 = 0.2
                    else:
                        amp1 = amp1
                    end1 = []
                    # print(y_end1_mo, y_end1_shi1, amp1)
                    # print(y_end_zzong)
                    y_find = y_end[0:366]
                    x_find = xx[0:366]
                    for ll in range(1, len(y_find)):
                        if y_find[ll] >= 0.1 and y_find[ll - 1] >= 0.1:
                            if y_find[ll - 1] <= y_find[ll]:
                                if y_find[ll - 1] < amp1:
                                    if y_find[ll] >= amp1:
                                        # print(1)
                                        k = (y_find[min(ll + 80, len(y_find) - 1)] - y_find[ll - 1]) / (
                                                x_find[min(ll + 80, len(x_find) - 1)] - x_find[ll - 1])
                                        kk = (y_find[min(ll + 50, len(y_find) - 1)] - y_find[ll - 1]) / (
                                                x_find[min(ll + 50, len(x_find) - 1)] - x_find[ll - 1])
                                        # print(lomax)
                                        if lomax < 100:
                                            lomax = 250
                                        else:
                                            lomax = lomax
                                        if k > 0 or kk > 0:
                                            if x_find[ll] < lomax:
                                                end1.append(x_find[ll])
                                        else:
                                            continue
                                    else:
                                        continue
                                else:
                                    continue
                            else:
                                continue
                    if 1 < len(end1):
                        gs[i, j, 0] = end1[-1]
                        # if im_geotrans0[3] + i * im_geotrans0[4] + j * im_geotrans0[5] > 20 and gs[i, j, 0] >= 305:
                        #     gs[i, j, 0] = numpy.random.randint(1, 10, 1)
                    elif len(end1) == 1:
                        gs[i, j, 0] = end1[0]
                    else:
                        gs[i, j, 0] = -99
                    # print(gs[i, j, 0])
                    # yy2.cell(1, 5).value = gs[i, j, 0]
                    if gs[i, j, 0] > 0:
                        st = gs[i, j, 0]
                        if gs[i, j, 0] < 200:
                            st = 200
                        else:
                            st = st
                        y_end3 = y_end[int(list(xx).index(st)):int(list(xx).index(gs[i, j, 0])) + 365]
                        xx_end3 = xx[int(list(xx).index(st)):int(list(xx).index(gs[i, j, 0])) + 365]
                        y_end3paixu = y_end3[numpy.argsort(y_end3[:])]
                        y_end3paixu = y_end3paixu[y_end3paixu != 0]
                        y_end1_shi2 = numpy.mean(y_end3paixu[0: math.floor(len(y_end3paixu) * 0.2)])
                        amplitude2 = y_end1_mo - y_end1_shi2
                        amp2 = amplitude2 * 0.2 + y_end1_shi2
                        # print(y_end1_mo, y_end1_shi2, amp2)
                        if amp2 > 0.2:
                            amp2 = amp2
                        else:
                            amp2 = 0.2
                        end2 = []
                        for ll1 in range(1, len(y_end3)):
                            if 0.1 <= y_end3[ll1] and y_end3[ll1 - 1] >= 0.1:
                                if y_end3[ll1 - 1] >= amp2:
                                    if y_end3[ll1] < amp2:
                                        k1 = (y_end3[min(ll1 + 20, len(y_end3) - 1)] - y_end[
                                            list(xx).index(gs[i, j, 0]) + ll1 - 20]) / (
                                                     xx_end3[min(ll1 + 20, len(y_end3) - 1)] - xx[
                                                 list(xx).index(gs[i, j, 0]) + ll1 - 20])
                                        if k1 < 0 and xx_end3[ll1 - 1] > xx[200] and xx_end3[ll1 - 1] - gs[
                                            i, j, 0] <= 365:
                                            end2.append(xx_end3[ll1 - 1])
                                        else:
                                            continue
                                    else:
                                        continue
                                else:
                                    continue
                            else:
                                continue
                        if len(end2) > 1:
                            # print(end2)
                            gs[i, j, 1] = end2[0]
                        elif len(end2) == 1:
                            gs[i, j, 1] = end2[0]
                        else:
                            gs[i, j, 1] = -99
                    else:
                        gs[i, j, 1] = -99
                    if gs[i, j, 1] >= 0 and gs[i, j, 0] >= 0:
                        gs[i, j, 2] = gs[i, j, 1] - gs[i, j, 0]

                    elif gs[i, j, 0] >= 0 and gs[i, j, 1] == -99:
                        gs[i, j, 2] = -99
                        # gs[i, j, 1] = gs[i, j, 2] + gs[i, j, 0]
                    else:
                        gs[i, j, 2] = -99
driver = gdal.GetDriverByName('Gtiff')
outRaster = driver.Create(outpath + '/' + 'growingseason_' + year + '.tif', im_width0, im_height0, 8,
                          gdal.GDT_Float32)
outRaster.SetGeoTransform(im_geotrans0)  # 参数2,6为水平垂直分辨率，参数3,5表示图片是指北的
outband1 = outRaster.GetRasterBand(1)
outband1.WriteArray(gs[:, :, 0], 0, 0)
outband2 = outRaster.GetRasterBand(2)
outband2.WriteArray(gs[:, :, 1], 0, 0)
outband3 = outRaster.GetRasterBand(3)
outband3.WriteArray(gs[:, :, 2], 0, 0)
outRaster.SetProjection(im_proj0)  #
outRaster.FlushCache()